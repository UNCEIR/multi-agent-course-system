# -*- coding: utf-8 -*-
"""成绩单 Excel 表头驱动解析 — 教师端 report 入参第一环（零 LLM）。

- sheet 选择：含「学号+姓名」表头且有数据行的 sheet（真实样本三 sheet 正确选中）
- 表头定位：含「姓名」单元格的行 = 表头首行；表头块 = 该行起 3 行（大类/合并延续/子类）
- 列模型：合并单元格感知（非只读模式 merged_cells.ranges）——
  子维度 = 「等级」的列 = 等级列；「分数/原始/折算/备注」列直接丢弃
- 元数据：行3「班级：X  学科：Y」提取，缺省回退文件名解析
- 值域：只取等级值，NFKC 归一；空值留空串不推断

Phase: 2 (implemented)
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from .contract import DROPPED_SUB_LABELS, ERR_PARSE_FAILED, is_valid_grade, normalize_grade

# 行3 元数据：班级：四（7）班  学科：道法   任课...
_META_RE = re.compile(r"班级[:：]\s*([^\s，,。]+).*?学科[:：]\s*([^\s，,。]+)")
# 文件名兜底：`（道法）四7班2023-2024第二学期成绩.xlsx`
_FILE_SUBJECT_RE = re.compile(r"[（(]([^（()）]{1,12})[）)]")
_FILE_CLASS_RE = re.compile(r"([一二三四五六]\d?班)")
_FILE_SEMESTER_RE = re.compile(r"(\d{4}-\d{4}第[一二两]学期|\d{4}-\d{4}学年)")


@dataclass
class ParsedStudent:
    student_id: str
    name: str
    grades: dict[str, str] = field(default_factory=dict)  # 子维度名 → 等级值


@dataclass
class ParsedFile:
    subject: str
    class_name: str
    semester: str
    source_name: str
    grade_columns: list[str]  # 维度名（按列序）
    students: list[ParsedStudent]
    warnings: list[str] = field(default_factory=list)


class ExcelParseError(Exception):
    """Excel 解析失败（带结构化原因，供上层转错误码）。"""

    def __init__(self, reason: str, *, file: str = "", sheet: str = "", row: int | None = None):
        self.reason = reason
        self.file = file
        self.sheet = sheet
        self.row = row
        super().__init__(f"[{ERR_PARSE_FAILED}] file={file} sheet={sheet} row={row}: {reason}")


def _select_data_sheet(wb) -> str:
    """选择数据 sheet：含「学号+姓名」表头且其后存在数据行的 sheet。

    真实样本：五（1）班（数据）、Sheet3（一年级模板空表）、Sheet1（空）。
    """
    for ws in wb.worksheets:
        header_row = _find_header_row(ws)
        if header_row is None:
            continue
        # 数据行判定：表头块后存在 学号列 非空的行
        for r in range(header_row + 3, min(ws.max_row + 1, header_row + 60)):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() != "":
                return ws.title
    return ""


def _find_header_row(ws) -> int | None:
    """定位含「姓名」单元格的表头行（扫描前 8 行）。"""
    for r in range(1, min(ws.max_row + 1, 9)):
        for c in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(row=r, column=c).value
            if v is not None and "姓名" in str(v):
                return r
    return None


def _merge_map(ws) -> dict[int, str]:
    """列号 → 父维度名（表头首行合并区间顶部单元格值）。

    对非合并列，父维度 = 该行单元格自身的值。标签统一折叠空白
    （真实样本 `综合性评价\n100分` → `综合性评价100分`，别名映射才可命中）。
    """
    header = _find_header_row(ws)
    if header is None:
        return {}
    mapping: dict[int, str] = {}

    def _norm(v) -> str:
        return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(v)))

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= header <= rng.max_row:
            value = ws.cell(row=rng.min_row, column=rng.min_col).value
            if value is not None:
                for c in range(rng.min_col, rng.max_col + 1):
                    mapping[c] = _norm(value)
    for c in range(1, ws.max_column + 1):
        if c not in mapping:
            v = ws.cell(row=header, column=c).value
            if v is not None and str(v).strip():
                mapping[c] = _norm(v)
    return mapping


def _sub_label(ws, header: int, col: int) -> str:
    """子维度标签：表头首行 +2（子类行）的值；空则 ""（单列维度）。"""
    r = header + 2
    if r > ws.max_row:
        return ""
    v = ws.cell(row=r, column=col).value
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(v))) if v is not None else ""


def _extract_meta(ws, filename: str) -> tuple[str, str, str]:
    """班级/学科/学期：行3 优先，文件名兜底。"""
    class_name, subject, semester = "", "", ""
    for r in range(1, min(ws.max_row + 1, 5)):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        m = _META_RE.search(str(v))
        if m:
            class_name, subject = m.group(1), m.group(2)
            break
    if not subject:
        m = _FILE_SUBJECT_RE.search(filename)
        if m:
            subject = m.group(1)
    if not class_name:
        m = _FILE_CLASS_RE.search(filename)
        if m:
            class_name = m.group(1)
    m = _FILE_SEMESTER_RE.search(filename)
    if m:
        semester = m.group(1)
    return class_name, subject, semester


def parse_workbook(path: str | Path) -> ParsedFile:
    """解析单个 Excel 工作簿 → ParsedFile（等级列驱动，丢分数/备注）。"""
    from openpyxl import load_workbook

    path = Path(path)
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ExcelParseError(f"无法打开文件: {exc}", file=path.name) from exc

    sheet_name = _select_data_sheet(wb)
    if not sheet_name:
        raise ExcelParseError("未找到含学号/姓名表头且有数据行的 sheet", file=path.name)
    ws = wb[sheet_name]

    header = _find_header_row(ws)
    parent_map = _merge_map(ws)
    class_name, subject, semester = _extract_meta(ws, path.name)

    # 构建等级列：子维度 == "等级"，或单列维度（子标签为空且父维度有效）
    grade_columns: list[str] = []
    grade_col_idx: list[int] = []
    for c in range(1, ws.max_column + 1):
        parent = parent_map.get(c, "")
        sub = _sub_label(ws, header, c)
        if parent in ("", "学号", "姓名", "备注"):
            continue
        if sub in DROPPED_SUB_LABELS:
            continue
        if sub == "等级" or sub == "":
            dim = parent if sub == "" else f"{parent}·{sub}"
            grade_columns.append(dim)
            grade_col_idx.append(c)
    if not grade_columns:
        raise ExcelParseError("未识别到任何等级列（表头结构不兼容）", file=path.name, sheet=sheet_name)

    # 数据行：表头块（header..header+2）之后首个非空行起
    start = None
    for r in range(header + 3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            start = r
            break
    if start is None:
        raise ExcelParseError("表头后无数据行", file=path.name, sheet=sheet_name)

    students: list[ParsedStudent] = []
    warnings: list[str] = []
    for r in range(start, ws.max_row + 1):
        sid = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if sid is None and name is None:
            continue
        if name is None:
            warnings.append(f"row {r}: 缺姓名，跳过")
            continue
        sid_norm = unicodedata.normalize("NFKC", str(sid)).strip() if sid is not None else ""
        grades: dict[str, str] = {}
        for idx, c in enumerate(grade_col_idx):
            raw = ws.cell(row=r, column=c).value
            grade = normalize_grade(raw)
            if not is_valid_grade(grade):
                warnings.append(f"row {r} {name}: {grade_columns[idx]} 等级非法 '{grade}'，置空")
                grade = ""
            grades[grade_columns[idx]] = grade
        students.append(ParsedStudent(student_id=sid_norm, name=str(name).strip(), grades=grades))

    return ParsedFile(
        subject=subject,
        class_name=class_name,
        semester=semester,
        source_name=path.name,
        grade_columns=grade_columns,
        students=students,
        warnings=warnings,
    )


def parse_score_excels(file_paths: list[str]) -> dict:
    """@tool 用入口：多文件解析 → 结构化结果（file_key → ParsedFile 序列化）。"""
    results = []
    for fp in file_paths:
        pf = parse_workbook(fp)
        results.append(
            {
                "file": pf.source_name,
                "subject": pf.subject,
                "class_name": pf.class_name,
                "semester": pf.semester,
                "grade_columns": pf.grade_columns,
                "students": [
                    {"student_id": s.student_id, "name": s.name, "grades": s.grades}
                    for s in pf.students
                ],
                "warnings": pf.warnings,
            }
        )
    return {"files": results}
