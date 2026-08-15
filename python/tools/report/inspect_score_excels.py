# -*- coding: utf-8 -*-
"""成绩单文件摘要工具 — 分类决策节点输入（不解析数据，零 LLM）。

读取每文件的元信息：学科/班级/是否含道法列/是否含「必选-自选」特征列/学生数。
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

from langchain_core.tools import tool
from pydantic import BaseModel


class InspectScoreExcelsInput(BaseModel):
    """inspect_score_excels 工具输入参数（file_keys 从请求上下文注入）。"""

    pass  # noqa: PIE790


def _header_texts(path: Path) -> list[str]:
    """读取表头区文本（前 6 行去空白，用于特征判定）。"""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return []
    texts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for v in row:
                if v is not None:
                    texts.append(unicodedata.normalize("NFKC", str(v)).strip())
    return texts


def _count_data_rows(path: Path) -> int:
    """轻量统计学生行数（表头后非空行）。"""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return 0
    total = 0
    for ws in wb.worksheets:
        # 找「姓名」表头行
        header = None
        for r in range(1, min(ws.max_row + 1, 9)):
            for c in range(1, min(ws.max_column + 1, 12)):
                v = ws.cell(row=r, column=c).value
                if v is not None and "姓名" in str(v):
                    header = r
                    break
            if header:
                break
        if header is None:
            continue
        for r in range(header + 3, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip():
                total += 1
    return total


@tool(args_schema=InspectScoreExcelsInput)
def inspect_score_excels() -> dict:
    """检查成绩单 Excel 文件摘要（学科/班级/特征列/学生数），供年级分类决策。

    file_keys 由请求上下文注入，无需参数。
    """
    from tools.report.render_report_batch import report_files_ctx

    file_keys = list(report_files_ctx.get() or [])
    if not file_keys:
        return {"files": [], "note": "未收到文件"}

    files = []
    for fp in file_keys:
        path = Path(fp)
        if not path.is_file():
            files.append({"file": path.name, "error": "file_not_found"})
            continue
        texts = _header_texts(path)
        joined = " ".join(texts)
        has_daofa = any("道法" in t or "道德与法治" in t for t in texts)
        has_required_optional = any("必选" in t or "自选" in t for t in texts)
        class_name = ""
        subject = ""
        for t in texts[:6]:
            if "班级" in t:
                for part in t.split():
                    if "班级" in part:
                        class_name = part.split("：")[-1].split(":")[-1].strip()
                        break
            if "学科" in t:
                for part in t.split():
                    if "学科" in part:
                        subject = part.split("：")[-1].split(":")[-1].strip()
                        break
        files.append(
            {
                "file": path.name,
                "subject": subject,
                "class_name": class_name,
                "has_daofa": has_daofa,
                "has_required_optional": has_required_optional,
                "student_count": _count_data_rows(path),
            }
        )
    return {"files": files, "file_count": len(files)}
